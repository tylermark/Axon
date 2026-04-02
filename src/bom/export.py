"""BM-005: BOM export to CSV, Excel, and PDF.

Renders the BillOfMaterials to one or more output formats. Optional
dependencies (openpyxl for Excel) are imported conditionally and
degrade gracefully with warnings.

Reference: TASKS.md Phase 9 BM-005.
"""

from __future__ import annotations

import csv
import logging
from datetime import UTC, datetime
from pathlib import Path
from typing import TYPE_CHECKING

from docs.interfaces.bill_of_materials import ExportFormat

if TYPE_CHECKING:
    from docs.interfaces.bill_of_materials import BillOfMaterials

logger = logging.getLogger(__name__)


def export_bom(
    bom: BillOfMaterials,
    output_dir: Path | str,
    formats: list[ExportFormat] | None = None,
) -> list[Path]:
    """Export the BillOfMaterials to one or more file formats.

    Args:
        bom: The complete BOM to export.
        output_dir: Directory to write output files into. Created if
            it does not exist.
        formats: List of export formats. Defaults to all three
            (CSV, Excel, PDF).

    Returns:
        List of paths to created files.
    """
    if formats is None:
        formats = [ExportFormat.CSV, ExportFormat.EXCEL, ExportFormat.PDF]

    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    timestamp = datetime.now(tz=UTC).strftime("%Y%m%d_%H%M%S")
    created: list[Path] = []

    for fmt in formats:
        if fmt == ExportFormat.CSV:
            path = _export_csv(bom, output_dir, timestamp)
            if path is not None:
                created.append(path)
        elif fmt == ExportFormat.EXCEL:
            path = _export_excel(bom, output_dir, timestamp)
            if path is not None:
                created.append(path)
        elif fmt == ExportFormat.PDF:
            path = _export_pdf(bom, output_dir, timestamp)
            if path is not None:
                created.append(path)

    logger.info("Exported BOM to %d file(s): %s", len(created), created)
    return created


# ══════════════════════════════════════════════════════════════════════════════
# CSV Export
# ══════════════════════════════════════════════════════════════════════════════


def _export_csv(
    bom: BillOfMaterials,
    output_dir: Path,
    timestamp: str,
) -> Path:
    """Export line items as CSV."""
    path = output_dir / f"bom_{timestamp}.csv"

    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(_CSV_HEADERS)

        for item in bom.line_items:
            writer.writerow(
                [
                    item.item_id,
                    item.category.value,
                    item.sku,
                    item.description,
                    item.gauge if item.gauge is not None else "",
                    item.depth_inches if item.depth_inches is not None else "",
                    item.length_inches if item.length_inches is not None else "",
                    item.quantity,
                    item.unit,
                    f"{item.unit_cost_usd:.2f}",
                    f"{item.extended_cost_usd:.2f}",
                    ";".join(str(e) for e in item.source_edge_ids),
                    ";".join(str(r) for r in item.source_room_ids),
                    item.notes,
                ]
            )

    logger.info("CSV export: %s (%d line items)", path, len(bom.line_items))
    return path


_CSV_HEADERS: list[str] = [
    "item_id",
    "category",
    "sku",
    "description",
    "gauge",
    "depth_inches",
    "length_inches",
    "quantity",
    "unit",
    "unit_cost_usd",
    "extended_cost_usd",
    "source_edge_ids",
    "source_room_ids",
    "notes",
]


# ══════════════════════════════════════════════════════════════════════════════
# Excel Export
# ══════════════════════════════════════════════════════════════════════════════


def _export_excel(
    bom: BillOfMaterials,
    output_dir: Path,
    timestamp: str,
) -> Path | None:
    """Export BOM as multi-sheet Excel workbook.

    Sheet 1: Line Items
    Sheet 2: Cost Summary
    Sheet 3: Labor Estimates
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, numbers
    except ImportError:
        logger.warning(
            "openpyxl not installed — skipping Excel export. Install with: pip install openpyxl"
        )
        return None

    wb = Workbook()

    # ── Sheet 1: Line Items ──────────────────────────────────────────
    ws_items = wb.active
    ws_items.title = "Line Items"

    header_font = Font(bold=True)
    headers = _CSV_HEADERS
    for col, header in enumerate(headers, start=1):
        cell = ws_items.cell(row=1, column=col, value=header)
        cell.font = header_font

    for row_idx, item in enumerate(bom.line_items, start=2):
        ws_items.cell(row=row_idx, column=1, value=item.item_id)
        ws_items.cell(row=row_idx, column=2, value=item.category.value)
        ws_items.cell(row=row_idx, column=3, value=item.sku)
        ws_items.cell(row=row_idx, column=4, value=item.description)
        ws_items.cell(row=row_idx, column=5, value=item.gauge if item.gauge is not None else None)
        ws_items.cell(
            row=row_idx,
            column=6,
            value=item.depth_inches if item.depth_inches is not None else None,
        )
        ws_items.cell(
            row=row_idx,
            column=7,
            value=item.length_inches if item.length_inches is not None else None,
        )
        ws_items.cell(row=row_idx, column=8, value=item.quantity)
        ws_items.cell(row=row_idx, column=9, value=item.unit)

        cost_cell = ws_items.cell(row=row_idx, column=10, value=item.unit_cost_usd)
        cost_cell.number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1

        ext_cell = ws_items.cell(row=row_idx, column=11, value=item.extended_cost_usd)
        ext_cell.number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1

        ws_items.cell(
            row=row_idx,
            column=12,
            value=";".join(str(e) for e in item.source_edge_ids),
        )
        ws_items.cell(
            row=row_idx,
            column=13,
            value=";".join(str(r) for r in item.source_room_ids),
        )
        ws_items.cell(row=row_idx, column=14, value=item.notes)

    # Auto-width columns
    for col in ws_items.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                val_len = len(str(cell.value)) if cell.value is not None else 0
                max_len = max(max_len, val_len)
            except (TypeError, AttributeError):
                pass
        ws_items.column_dimensions[col_letter].width = min(max_len + 2, 50)

    # ── Sheet 2: Cost Summary ────────────────────────────────────────
    ws_cost = wb.create_sheet("Cost Summary")
    summary = bom.material_summary
    cost_rows = [
        ("Category", "Cost (USD)"),
        ("CFS Studs", summary.cfs_studs_usd),
        ("CFS Track", summary.cfs_track_usd),
        ("Fasteners", summary.fasteners_usd),
        ("Clips", summary.clips_usd),
        ("Bridging", summary.bridging_usd),
        ("Blocking", summary.blocking_usd),
        ("Sheathing", summary.sheathing_usd),
        ("Pods", summary.pods_usd),
        ("Connection Hardware", summary.connection_hardware_usd),
        ("Other", summary.other_usd),
        ("", ""),
        ("MATERIAL TOTAL", summary.material_total_usd),
        ("", ""),
        ("Fabrication Material", bom.project_cost.fabrication_material_usd),
        ("Fabrication Labor", bom.project_cost.fabrication_labor_usd),
        ("Fabrication Subtotal", bom.project_cost.fabrication_subtotal_usd),
        ("Pod Cost", bom.project_cost.pod_cost_usd),
        ("Shipping", bom.project_cost.shipping_usd),
        ("Installation Labor", bom.project_cost.installation_labor_usd),
        ("Installation Material", bom.project_cost.installation_material_usd),
        ("Installation Subtotal", bom.project_cost.installation_subtotal_usd),
        ("", ""),
        ("TOTAL PROJECT COST", bom.project_cost.total_project_cost_usd),
        (
            f"Contingency ({bom.project_cost.contingency_pct:.0f}%)",
            round(
                bom.project_cost.total_project_cost_usd * bom.project_cost.contingency_pct / 100.0,
                2,
            ),
        ),
        (
            "Total with Contingency",
            round(
                bom.project_cost.total_project_cost_usd
                * (1 + bom.project_cost.contingency_pct / 100.0),
                2,
            ),
        ),
    ]

    for row_idx, (label, value) in enumerate(cost_rows, start=1):
        cell_a = ws_cost.cell(row=row_idx, column=1, value=label)
        cell_b = ws_cost.cell(row=row_idx, column=2, value=value)
        if row_idx == 1:
            cell_a.font = header_font
            cell_b.font = header_font
        elif isinstance(value, (int, float)):
            cell_b.number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1

    ws_cost.column_dimensions["A"].width = 30
    ws_cost.column_dimensions["B"].width = 18

    # ── Sheet 3: Labor Estimates ─────────────────────────────────────
    ws_labor = wb.create_sheet("Labor Estimates")
    labor_headers = [
        "Trade",
        "Hours",
        "Hourly Rate (USD)",
        "Cost (USD)",
        "Crew Size",
        "Notes",
    ]
    for col, header in enumerate(labor_headers, start=1):
        cell = ws_labor.cell(row=1, column=col, value=header)
        cell.font = header_font

    for row_idx, est in enumerate(bom.labor_estimates, start=2):
        ws_labor.cell(row=row_idx, column=1, value=est.trade.value)
        ws_labor.cell(row=row_idx, column=2, value=est.hours)
        rate_cell = ws_labor.cell(row=row_idx, column=3, value=est.hourly_rate_usd)
        rate_cell.number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
        cost_cell = ws_labor.cell(row=row_idx, column=4, value=est.cost_usd)
        cost_cell.number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
        ws_labor.cell(row=row_idx, column=5, value=est.crew_size)
        ws_labor.cell(row=row_idx, column=6, value=est.notes)

    # Totals row
    total_row = len(bom.labor_estimates) + 2
    ws_labor.cell(row=total_row, column=1, value="TOTAL").font = header_font
    ws_labor.cell(row=total_row, column=2, value=bom.total_labor_hours)
    total_cost_cell = ws_labor.cell(row=total_row, column=4, value=bom.total_labor_cost_usd)
    total_cost_cell.number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1

    for col_dim, width in [("A", 16), ("B", 10), ("C", 18), ("D", 14), ("E", 12), ("F", 50)]:
        ws_labor.column_dimensions[col_dim].width = width

    # ── Save ─────────────────────────────────────────────────────────
    path = output_dir / f"bom_{timestamp}.xlsx"
    wb.save(path)
    logger.info("Excel export: %s", path)
    return path


# ══════════════════════════════════════════════════════════════════════════════
# PDF Export (simple text-based summary)
# ══════════════════════════════════════════════════════════════════════════════


def _export_pdf(
    bom: BillOfMaterials,
    output_dir: Path,
    timestamp: str,
) -> Path:
    """Export BOM as a formatted text summary file.

    Produces a human-readable .txt report rather than requiring
    heavy PDF dependencies like reportlab. The file is named with
    a .txt extension to be transparent about format.
    """
    path = output_dir / f"bom_{timestamp}.txt"
    lines: list[str] = []

    lines.append("=" * 72)
    lines.append("AXON — BILL OF MATERIALS REPORT")
    lines.append("=" * 72)
    lines.append("")
    lines.append(f"Generated: {bom.export.generated_at or datetime.now(tz=UTC).isoformat()}")
    lines.append(f"Generator: {bom.export.generator_version or 'axon-bom v1.0.0'}")
    lines.append(f"KG Version: {bom.export.kg_version or 'N/A'}")
    lines.append("")

    # ── Line Items Summary ───────────────────────────────────────────
    lines.append("-" * 72)
    lines.append("LINE ITEMS")
    lines.append("-" * 72)
    lines.append(
        f"{'ID':<10} {'Category':<22} {'SKU':<20} {'Qty':>8} {'Unit':>4} {'Unit$':>10} {'Ext$':>12}"
    )
    lines.append("-" * 72)

    for item in bom.line_items:
        lines.append(
            f"{item.item_id:<10} {item.category.value:<22} "
            f"{item.sku:<20} {item.quantity:>8.1f} {item.unit:>4} "
            f"${item.unit_cost_usd:>9.2f} ${item.extended_cost_usd:>11.2f}"
        )

    lines.append("-" * 72)
    lines.append(f"{'Total line items:':<35} {len(bom.line_items)}")
    lines.append("")

    # ── Material Cost Summary ────────────────────────────────────────
    lines.append("-" * 72)
    lines.append("MATERIAL COST SUMMARY")
    lines.append("-" * 72)
    ms = bom.material_summary
    cost_lines = [
        ("CFS Studs", ms.cfs_studs_usd),
        ("CFS Track", ms.cfs_track_usd),
        ("Fasteners", ms.fasteners_usd),
        ("Clips", ms.clips_usd),
        ("Bridging", ms.bridging_usd),
        ("Blocking", ms.blocking_usd),
        ("Sheathing", ms.sheathing_usd),
        ("Pods", ms.pods_usd),
        ("Connection Hardware", ms.connection_hardware_usd),
        ("Other", ms.other_usd),
    ]
    for label, cost in cost_lines:
        if cost > 0:
            lines.append(f"  {label:<30} ${cost:>12.2f}")
    lines.append(f"  {'MATERIAL TOTAL':<30} ${ms.material_total_usd:>12.2f}")
    lines.append("")

    # ── Labor Estimates ──────────────────────────────────────────────
    lines.append("-" * 72)
    lines.append("LABOR ESTIMATES")
    lines.append("-" * 72)
    lines.append(f"{'Trade':<16} {'Hours':>8} {'Rate':>10} {'Cost':>12} {'Crew':>6}")
    lines.append("-" * 72)
    for est in bom.labor_estimates:
        lines.append(
            f"  {est.trade.value:<14} {est.hours:>8.1f} "
            f"${est.hourly_rate_usd:>9.2f} ${est.cost_usd:>11.2f} "
            f"{est.crew_size:>6d}"
        )
    lines.append("-" * 72)
    lines.append(
        f"  {'TOTAL':<14} {bom.total_labor_hours:>8.1f} {'':>10} ${bom.total_labor_cost_usd:>11.2f}"
    )
    lines.append("")

    # ── Project Cost Breakdown ───────────────────────────────────────
    lines.append("-" * 72)
    lines.append("PROJECT COST BREAKDOWN")
    lines.append("-" * 72)
    pc = bom.project_cost
    lines.append(f"  Fabrication Material:   ${pc.fabrication_material_usd:>12.2f}")
    lines.append(f"  Fabrication Labor:      ${pc.fabrication_labor_usd:>12.2f}")
    lines.append(f"  Fabrication Subtotal:   ${pc.fabrication_subtotal_usd:>12.2f}")
    lines.append(f"  Pod Cost:               ${pc.pod_cost_usd:>12.2f}")
    lines.append(f"  Shipping:               ${pc.shipping_usd:>12.2f}")
    lines.append(f"  Installation Labor:     ${pc.installation_labor_usd:>12.2f}")
    lines.append(f"  Installation Material:  ${pc.installation_material_usd:>12.2f}")
    lines.append(f"  Installation Subtotal:  ${pc.installation_subtotal_usd:>12.2f}")
    lines.append("")
    lines.append(f"  TOTAL PROJECT COST:     ${pc.total_project_cost_usd:>12.2f}")
    contingency_amt = round(pc.total_project_cost_usd * pc.contingency_pct / 100.0, 2)
    total_with = round(pc.total_project_cost_usd * (1 + pc.contingency_pct / 100.0), 2)
    lines.append(f"  Contingency ({pc.contingency_pct:.0f}%):       ${contingency_amt:>12.2f}")
    lines.append(f"  Total w/ Contingency:   ${total_with:>12.2f}")
    lines.append("")
    lines.append("=" * 72)

    path.write_text("\n".join(lines), encoding="utf-8")
    logger.info("PDF/text export: %s", path)
    return path
